"""
Cells
 for Company Name : d13
 Company/client street adr : d14
 Cit,state = d15
 zip = d16
 phone num = d17

 IF I DECIDE TO IMPLEMEMNT THIS SECTION
 comments
    cells e14 -> e17

 BASE BID
    DATE = B21
    DESC = C21
    AMOUNT = E21

 Addition costs
    DATE = B22
    DESC = C22
    AMOUNT = E22

 total cost
    e35
 Inc the row number to add additional costs if they exists


Example of loading a workbook
https://stackoverflow.com/questions/29355778/openpyxl-load-the-workbook-and-save-open-saved-file-with-error-message


wb.save('document_template.xltx', as_template=False)

"""
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
from pathlib import Path

path_to_template = "D:\Capstone\Capstone_FX\ExcelFiles\Template\JASWInvoiceTemplate.xlsx"
bid_cells = ["B21", "C21", "E21"]
client_info_cells = ["D13", "D14",  "D15", "D16", "D17"]
additional_cost_cell = ["B22", "C22", "E22",
                        "B23", "C23", "E23",
                        "B24", "C24", "E24",
                        "B25", "C25", "E25",
                        "B26", "C26", "E26",
                        "B27", "C27", "E27",
                        "B28", "C28", "E28"]
total_cell = "E35"
"""
First Extract the information from the file to fill the excel file
"""
home = "D:\Capstone\Capstone_FX\ExcelFiles\DataFiles\DataFile.txt"
data_file = open(home, "r")


data_array = []

data_array = data_file.read().split()

data_file.close()

index = 0
customer_info_array = []
customer_address_street = []
customer_address = []
customer_bid = 0;
customer_additonal_info = []
customer_company = []
print(data_array)

for element in data_array:
    if(element == "End"):
        index += 1
    else:
        if(index == 0):
            customer_info_array.append(element)
        elif(index == 2):
            customer_address_street.append(element)
        elif (index == 3):
            customer_address.append(element)
        elif(index ==4):
            customer_additonal_info.append(element)
        elif(index==5):
            customer_bid = element
        elif (index == 1):
            customer_company.append(element)

print(customer_info_array)
print(customer_company)
print(customer_address_street)
print(customer_address)
print(customer_additonal_info)
print(customer_bid)

total_cost = 0

wb = load_workbook(path_to_template)
ws = wb.active
ws[bid_cells[0]] = datetime.date.today()
ws[bid_cells[1]] = ""
ws[bid_cells[2]] = customer_bid

if(customer_company[0] != "NA"):
    company = ""
    for word in customer_company:
        company += word + ' '

    ws[client_info_cells[0]] = company
else:
    ws[client_info_cells[0]] = customer_info_array[0] + " " + customer_info_array[1]

street_address = ""
for element in customer_address_street:
    street_address += element + " "

ws[client_info_cells[1]] = street_address

ws[client_info_cells[2]] = customer_address[0] + " " + customer_address[1]
ws[client_info_cells[3]] = customer_address[2]

ws[client_info_cells[4]] = customer_info_array[2]

additional = ""
multiple_lines = 0

for element in customer_additonal_info:
    additional += element + " "

additional = additional.split(',')

length_of_array = len(additional_cost_cell)

index=0
for x in range(0, length_of_array, 3):
    if index > len(additional)-2:
        break
    ws[additional_cost_cell[x]] = datetime.date.today()
    ws[additional_cost_cell[x+1]] = additional[index]
    ws[additional_cost_cell[x+2]] = additional[index+1]
    index +=2


total_cost += float(customer_bid)

for x in range(1, len(additional),2):
    total_cost += float(additional[x])

total_cost = str(total_cost)
ws[total_cell] = total_cost


cust_name =  customer_info_array[0] + " " + customer_info_array[1]
path_to_save = "D:\Capstone\Capstone_FX\ExcelFiles\CustomerFiles\\"+cust_name
name_of_file = str(datetime.date.today())
wb.save(path_to_save+name_of_file + ".xlsx")





